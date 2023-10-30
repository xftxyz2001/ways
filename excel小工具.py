import threading
from selenium import webdriver
from openpyxl import load_workbook
import os
import tkinter as tk


# 获取driver
def get_driver():
    try:
        return webdriver.Edge()
    except:
        print("浏览器启动失败，请检查浏览器驱动：https://developer.microsoft.com/zh-cn/microsoft-edge/tools/webdriver/")


# 采用前缀的方式获取需要打开的文件名
def get_filename():
    fn = input("请输入文件名（或前缀）：")
    if fn is None or fn == "":
        print("文件名为空，请重新输入")
        return get_filename()

    for dirpath, dirnames, filenames in os.walk('.'):
        for filename in filenames:
            if filename.startswith(fn):
                fn = filename
                break

    print("获取到文件："+fn)
    # xlsx/xlsm/xltx/xltm
    if fn.endswith(".xlsx") or fn.endswith(".xlsm") or fn.endswith(".xltx") or fn.endswith(".xltm"):
        return fn
    else:
        print("文件类型错误，请输入xlsx/xlsm/xltx/xltm文件")
        return get_filename()


# 获取一个工作表的第一行
def get_header(ws):
    header = []
    for rowheader in ws.iter_rows(min_row=1, max_row=1):
        for cell in rowheader:
            header.append(cell.value)
    print("获取到表头："+str(header))
    return header


fn = get_filename()  # 获取文件名
browser = get_driver()  # 获取浏览器驱动
wb = load_workbook(fn)  # 载入文件
ws = wb.active  # 获取当前活动的工作表
headers = get_header(ws)  # 获取表头
root_window = tk.Tk()  # 主窗口
texts = []  # 文本监听
current_row = 2  # 当前操作行
jump_text = tk.Text(root_window, width=25, height=1)  # 跳转行文本框
is_submit = False  # 是否提交过


def browser_get(url):
    global browser
    try:
        _ = browser.window_handles
    except Exception as _:
        print("检测到浏览器异常关闭，正在重新打开")
        browser = get_driver()
    finally:
        try:
            print("请求页面："+url)
            browser.get(url)
            print("请求成功")
        except:
            print("页面请求失败")


def print_max_cell(r):
    # 获取最长的一条数据，打印到控制台参考
    max_content = ""
    max_len = 0
    for i, v in enumerate(r):
        if len(v) > max_len:
            max_len = len(v)
            max_content = v
    print(str(max_len) + "<"*100)
    print(max_content)
    print(">"*100)


# 载入新的行
def load_strvars():
    global current_row, texts, wb, ws, root_window, browser
    root_window.title("正在载入第"+str(current_row)+"行，共"+str(ws.max_row)+"行")
    row = []  # 存放当前行的数据
    # 获取当前行的数据
    for rowheader in ws.iter_rows(min_row=current_row, max_row=current_row):
        for cell in rowheader:
            v = cell.value
            if v is None or str(v).strip() == '*':
                v = ""
            v = str(v).strip()
            row.append(v)
    threading.Thread(target=print_max_cell, args=(row,)).start()
    # 渲染数据
    for i, strvar in enumerate(texts):
        strvar.delete(1.0, "end")
        strvar.insert("end", row[i])
    # browser_get(row[0])
    threading.Thread(target=browser_get, args=(row[0],)).start()
    root_window.title("正在编辑第"+str(current_row)+"行，共"+str(ws.max_row)+"行")


def up():
    global current_row
    current_row -= 1
    if (current_row < 2):
        print("行号越界，已定位到第2行")
        current_row = 2
    load_strvars()


def down():
    global current_row
    current_row += 1
    if (current_row > ws.max_row):
        print("行号越界，已定位到最后一行")
        current_row = ws.max_row
    load_strvars()


def jump():
    global current_row
    try:
        current_row = int(jump_text.get(1.0, "end").strip())
        if (current_row < 2):
            print("行号越界，已定位到第2行")
            current_row = 2
        if (current_row > ws.max_row):
            print("行号越界，已定位到最后一行")
            current_row = ws.max_row
        load_strvars()
    except:
        print("非法行号")
    jump_text.delete(1.0, "end")


def save():
    global wb, fn, is_submit
    if not is_submit:
        print("自上次保存以来，未发生过提交")
        return
    wb.save(fn)
    print("保存成功！")
    is_submit = False


def window_close():
    global root_window
    print("检测到窗口关闭，正在保存文件")
    save()
    root_window.destroy()


# 提交
def submit():
    global current_row, texts, root_window, is_submit
    root_window.title("正在提交第"+str(current_row)+"行，共"+str(ws.max_row)+"行")
    is_submit = True
    strs = []
    for v in texts:
        s = v.get(1.0, "end")
        if s is None or len(str(s).strip()) <= 0:
            s = "*"
        strs.append(s)
    for i, s in enumerate(strs):
        ws.cell(row=current_row, column=i+1).value = s


def submit2down():
    submit()
    down()


# 创建标签和输入框并处理布局
for i, header in enumerate(headers):
    tk.Label(root_window, text=header).grid(
        row=i//4, column=(i % 4) * 2)
    texts.append(tk.Text(root_window, width=25, height=3, undo=True))
    texts[-1].grid(row=i//4, column=(i % 4) * 2 + 1)

tk.Button(root_window, text="上一行", command=up).grid(
    row=i//4+1, column=0)
tk.Button(root_window, text="下一行", command=down).grid(
    row=i//4+1, column=1)
tk.Button(root_window, text="提交当前行", command=submit).grid(
    row=i//4+1, column=2)
tk.Button(root_window, text="提交并切换到下一行", command=submit2down).grid(
    row=i//4+1, column=3)
tk.Button(root_window, text="保存已提交行", command=save).grid(
    row=i//4+1, column=4)
jump_text.grid(row=i//4+1, column=5)
tk.Button(root_window, text="跳转", command=jump).grid(
    row=i//4+1, column=6)


load_strvars()


# 启动配置
os.system('title '+fn+' - 编辑中'+"     By xftxyz")
print("初始化已完成。")
root_window.protocol("WM_DELETE_WINDOW", window_close)
root_window.mainloop()
if browser is not None:
    browser.quit()
print("程序终止！")